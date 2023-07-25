from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, View
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from keras.models import load_model
from sklearn.preprocessing import StandardScaler

from smart_home_monitoring.forms import UploadDataForm
from .converters import ResultSummary, OrderedLabelEncoder, TextLabelEncoderDummy

# import xlwt





def index(request):
    # template_name = "index.html"
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        print("\nUser Name = ",username)
        print("Password = ",password)
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('smart_home_monitoring:dashboard')
            
        else:
            context = {'message':'Invalid User Name and Password'}
            return render(request, 'index.html', context)
    return render(request, 'index.html', {'name': 'admin', 'pass': 'Info@123'})

        

def logoutUser(request):
    logout(request)
    return redirect('home')


class DashboardView(LoginRequiredMixin,TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        # context['cls'] = Lecturer.objects.count()
        # context['results'] = Results.objects.count()
        # context['students'] = Student.objects.count()
        # context['subjects'] = Subject.objects.count()
        return context
    

# Build paths inside the project like this: BASE_DIR / 'subdir'.
# BASE_DIR = Path(__file__).resolve().parent.parent
directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    

model = os.path.join(directory, 'smart_home_monitoring/smart_home/smart_home_model.h5')
# print(model)
activity = pd.read_csv(os.path.join(directory, 'smart_home_monitoring/smart_home/activity.csv'), index_col=False) 


def upload_data(request):
    context = {}
    if request.method == "POST":
        form = UploadDataForm(request.POST, request.FILES)
        print(request.FILES.get('excel_file'))
        if form.is_valid():
            file = request.FILES.get('excel_file')
            pred_range = request.POST['pred_range']
        
            wb = load_workbook(file, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            df = pd.DataFrame(sheet.values)

            df.columns = df.iloc[0]
            yy_train_df = activity #df['activity']
            yy = df['activity']
            df = df.iloc[1:,]
            print(yy_train_df)

            X_test_df = df.drop(labels=['numDistinctSensors','sensorCount-Chair','sensorCount-DiningRoom',
             'sensorCount-Hall','sensorCount-Office','sensorCount-WorkArea','sensorElTime-Chair','sensorElTime-DiningRoom',
             'sensorElTime-Hall','sensorElTime-Office',
             'sensorElTime-WorkArea','lastSensorLocation','lastMotionLocation', 'sensorElTime-OutsideDoor', 'lastSensorEventSeconds','activity'], axis=1)

            X_test_df = X_test_df.to_numpy() # test
            # Scaling data
            sc = StandardScaler()
            # ss_train = sc.fit_transform(xx_train_df) # train
            S_test = sc.fit_transform(X_test_df)
            # s_val_pred = sc.transform(ss_val_df)
            # print(S_test)

            encoded_Y_pred, encoder = TextLabelEncoderDummy.labelencoder(yy_train_df)
            dummy_y_pred, uniques = TextLabelEncoderDummy.encoded_to_dummy(encoded_Y_pred)
            print('uniques '+str(uniques))
            saved_model = load_model(model)
            # saved_model.summary()

            # 5. make predictions
            pred = saved_model.predict(S_test, verbose=1)
            print(pred.shape, uniques.shape)

            # Reverse predictios
            reverse_dummy_pred =TextLabelEncoderDummy.reverse_dummy_to_encoded(pred,uniques)
            print(pd.unique(reverse_dummy_pred).shape)

            reverse_encoded_y_pred = TextLabelEncoderDummy.reverse_encoded_to_text(reverse_dummy_pred,encoder)
            print(reverse_encoded_y_pred.shape)
            yy=yy.iloc[1:,]
            object_list = zip(reverse_encoded_y_pred,yy)
            print(yy.shape)

            field_list = ['Actual', 'Predicted']
                
            context={'panel_title':'Smart Home Simulator',
            'field_list':field_list,
            'object_list':object_list}
            return render(request, "smart_home/smarthome_list.html", context)

        else:
            print(form.errors)
            # messages.success(request, 'Account was created for ' + username)
            return redirect('smart_home_monitoring:upload_data')
    else:
        form = UploadDataForm()
        context['main_page_title'] = 'Declare Students Result'
        context['panel_name'] = 'Results'
        context['panel_title'] = 'Declare Result'
        context['form'] = form
    return render(request, "smart_home/smarthome_list.html", context)



# directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    
file_db = os.path.join(directory, 'smart_home_monitoring/smart_home/smart_home_test_data.xlsx')

# cols = ['lastSensorEventHours', 'lastSensorEventSeconds', 'lastSensorDayOfWeek',
#     'windowDuration', 'timeSinceLastSensorEvent', 'prevDominantSensor1',
#     'prevDominantSensor2', 'lastSensorID', 'lastSensorLocation',
#     'lastMotionLocation', 'complexity', 'activityChange', 'areaTransitions',
#     'numDistinctSensors', 'sensorCount-Bathroom', 'sensorCount-Bedroom',
#     'sensorCount-Chair', 'sensorCount-DiningRoom', 'sensorCount-Hall',
#     'sensorCount-Ignore', 'sensorCount-Kitchen', 'sensorCount-LivingRoom',
#     'sensorCount-Office', 'sensorCount-OutsideDoor', 'sensorCount-WorkArea',
#     'sensorElTime-Bathroom', 'sensorElTime-Bedroom', 'sensorElTime-Chair',
#     'sensorElTime-DiningRoom', 'sensorElTime-Hall', 'sensorElTime-Ignore',
#     'sensorElTime-Kitchen', 'sensorElTime-LivingRoom',
#     'sensorElTime-Office', 'sensorElTime-OutsideDoor',
#     'sensorElTime-WorkArea','activities']


# def downloads(request):
# def downloads1(request):
#     response = HttpResponse(
#         content_type="application/ms-excel",
#         headers={"Content-Disposition": 'attachment; filename=smart home test data.xlsx'},
#     )
#     wb = xlwt.Workbook(encoding='utf-8')
#     ws = wb.add_sheet("sheet1")
#     row_num = 0
#     font_style = xlwt.XFStyle()
#     font_style.font.bold = True

#     for col_num in range(len(cols)):
#         ws.write(row_num, col_num, cols[col_num], font_style)
    
#     font_style = xlwt.XFStyle()
#     # data = file_db
#     # df = pd.read_csv(file_db)
#     wb1 = load_workbook(file_db, data_only=True)
#     sheet = wb1[wb1.sheetnames[0]]
#     df = pd.DataFrame(sheet.values)
#     print(df.values[0][0])
#     print(len(df))
#     print(len(df.values[0]))

#     for i in range(len(df)):
#         if i == (len(df)-1):
#             break
#         for d in range(len(df.values[i])):
#             ws.write(i+1,d,df.values[i+1][d])

#     wb.save(response)
#     return response


def downloads(request):
    response = HttpResponse(
        content_type="application/ms-excel",
        headers={"Content-Disposition": 'attachment; filename=smart home test data.xlsx'},
    )
    wb1 = load_workbook(file_db, data_only=True)
    sheet = wb1[wb1.sheetnames[0]]
    df = pd.DataFrame(sheet.values)

    wb = Workbook()
    dest_filename = 'empty_book.xlsx'
    # Get first sheet
    ws = wb.active
    # ws1.title = "Range"
    # ws = wb.create_sheet(title="Data")
    for row in range(len(df)):
        ws.append(list(df.values[row]))
    # for row in range(len(df)):
    #     print(df.values[row])
    #     for col in range(len(df.values[row])):
    #         # Fill the cell with the column letter
    #         # ws.cell(row=row+1, column=col+1, value="l")
    #         break
    
    # Save the workbook
    wb.save(response)
    return response


def download_page(request):
    # pass
    context = {}
    context['main_page_title'] = 'Download Smart Home Monitoring Records'
    context['panel_name'] = 'Download'
    context['panel_title'] = 'Download Smart Home Monitoring Test File'
    context['file'] = file_db
    return render(request, "smart_home/smarthome_download.html",context)


def about(request):
    # pass
    context = {}
    context['main_page_title'] = 'About Smart Home Monitoring Simulator'
    context['panel_name'] = 'About'
    context['panel_title'] = 'How to use the simulator'
    context['file'] = file_db
    return render(request, "smart_home/smarthome_about.html",context)






