from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, View
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from openpyxl import load_workbook
import os
import itertools
from django.utils import timezone

# from sklearn.preprocessing import StandardScaler
from court_cases_classification.models import UploadCourtCase
from court_cases_classification.forms import UploadCourtCaseForm
from .court_cases.predict import Predict
import pandas as pd
import csv

from helpers.decorators import unauthenticated_user
# from django.contrib.auth.decorators import login_required


def logoutUser(request):
    logout(request)
    return redirect('home')


class DashboardView(LoginRequiredMixin,TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        # context['cls'] = Lecturer.objects.count()
        # context['results'] = Results.objects.count()
        # context['students'] = Student.objects.count()
        # context['subjects'] = Subject.objects.count()
        return context
    

def index(request):
    # template_name = "index.html"
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        print("\nUser Name = ",username)
        print("Password = ",password)
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('smart_home_monitoring:dashboard')
            
        else:
            context = {'message':'Invalid User Name and Password'}
            return render(request, 'index.html', context)
    return render(request, 'index.html', {'name': 'admin', 'pass': 'Info@123'})

directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    
file_db = os.path.join(directory, 'court_cases_classification/court_cases/sherloc_court_cases_7.csv')


@unauthenticated_user
# @login_required
def upload_data(request):
    context = {}
    if request.method == "POST":
        form = UploadCourtCaseForm(request.POST, request.FILES)
        print(request.FILES.get('excel_file'))
        if form.is_valid():
            file = request.FILES.get('excel_file')
            text = request.POST['text']
            if file != None:
                # this will take from excel sheets
                wb = load_workbook(file, data_only=True)
                sheet = wb[wb.sheetnames[0]]
                df = pd.DataFrame(sheet.values,columns=['text'])
                # print(df)
                data = itertools.chain.from_iterable(df.values)
                # data = df.values
                # print(data)
            else:
                # transform data
                # text = pred_text.get('1.0','end')
                df = pd.DataFrame({'text':[text]})
                # print(df)
                data = itertools.chain.from_iterable(df.values)
                # data = df.values
            # print('dfhfhfhfh   ',df)
            dfff = df.copy()
            Xr = Predict.transform_text_r(dfff)
            print('dffffffffff   ',df)
            Xc = Predict.transform_text_c(df)
            # print(Xr)
            # print(Xc)
            predR = Predict.pred_r(Xr)
            predC = Predict.pred_c(Xc)

            predRList = []
            for i in predR:
                if i == 0:
                    predRList.append('Not Sentence')
                else:
                    predRList.append('Sentenced')

            # UploadCourtCaseForm.save(.)
            text = [i for i in data]
            # print(data)
            # print('*'*40)
            # print(text)
            # print('*'*40)
            # print(df)
            # print('*'*40)
            # print(predRList)
            # print('*'*40)
            # print(predC)
            # save record
            df1 = pd.DataFrame({'text':text,'crime_types':predC,'sentence':predRList})
            # df1 = pd.DataFrame({'crime_types':predC,'sentence':predRList})
            df2 = pd.read_csv(file_db)
            df3 = pd.concat([df2,df1],axis=0)
            df3.to_csv(file_db,index=False)

            # save result to db for safety
            # for text,sentence,crime in zip(data,predRList,predC):
            #     UploadCourtCase.objects.create(text=text, sentence=sentence, crime=crime)            
            # # pull db content and delete it later
            # result = UploadCourtCase.objects.filter(reg_date__lte=timezone.now()).values(
            # 'text','sentence','crime')
            # print(result)

            object_list = zip(text,predRList,predC)
            # object_list = result
            field_list = ['Court Case(s)','Sentence Status', 'Crime Type']
            context={'panel_title':'Court Cases Model',
            'field_list':field_list,
            'object_list':object_list}

            return render(request, "court_cases/court_cases_list.html", context)

        else:
            print(form.errors)
            # messages.success(request, 'Account was created for ' + username)
            return redirect('court_cases_classification:court_case')
    else:
        form = UploadCourtCaseForm()
        context['main_page_title'] = 'Check court case type and predict whether sentenced or not'
        context['panel_name'] = 'Court Case Model'
        context['panel_title'] = 'Predict'
        context['form'] = form
        return render(request, "court_cases/court_cases_list.html", context)



@unauthenticated_user
def downloads(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename=court cases.csv'},
    )
    df = pd.read_csv(file_db)
    writer = csv.writer(response)
    writer.writerow(["text", "crime_types", "sentence"])
    for i in range(len(df)):
        writer.writerow(df.values[i])
    return response

@unauthenticated_user
def download_page(request):
    # pass
    context = {}
    context['main_page_title'] = 'Download Court Cases Records'
    context['panel_name'] = 'Download'
    context['panel_title'] = 'Download Court Case File'
    context['file'] = file_db
    return render(request, "court_cases/court_cases_download.html",context)

@unauthenticated_user
def about(request):
    # pass
    context = {}
    context['main_page_title'] = 'About Court Cases Simulator'
    context['panel_name'] = 'About'
    context['panel_title'] = 'How to use the simulator'
    # context['file'] = file_db
    return render(request, "court_cases/court_cases_about.html",context)





